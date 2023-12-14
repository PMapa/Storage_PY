
from tkinter import ttk
import openpyxl
import pandas as pd
import tkinter as tk

def atualizar_table_frame():
    # Remover todos os widgets existentes no table_frame
    for widget in table_frame.winfo_children():
        widget.destroy()

    # Ler as informações do formulário
    quarto = quarto_entry.get()
    camareira = camareira_entry.get()
    itens_consumidos = itens_consumidos_entry.get()
    data = [quarto, camareira, itens_consumidos]
    
    # Adicionar as informações à planilha
    write_xlsx(data)
    
    # Limpar os campos do formulário
    quarto_entry.delete(0, tk.END)
    camareira_entry.delete(0, tk.END)
    itens_consumidos_entry.delete(0, tk.END)


    # Adicione seus novos dados aqui
    novos_dados = data

    # Recriar os widgets no table_frame com os novos dados
    for dado in novos_dados:
        # Supondo que você esteja usando Labels para exibir os dados
        label = tk.Label(table_frame, text=dado)
        label.pack()

def read_xlsx():
    # Abrir o arquivo xlsx
    wb = openpyxl.load_workbook('vendas.xlsx')
    # Selecionar a planilha
    ws = wb.active
    # Ler as informações das colunas A:F
    data = []
    for row in ws.iter_rows(min_row=0, max_col=6):
        data.append([cell.value for cell in row])
    return data

def write_xlsx(data):
    # Abrir o arquivo xlsx
    wb = openpyxl.load_workbook('vendas.xlsx')
    # Selecionar a planilha
    ws = wb.active
    # Adicionar uma nova linha com as informações do formulário
    ws.append(data)
    # Salvar o arquivo xlsx
    wb.save('vendas.xlsx')

# Criar a janela principal
root = tk.Tk()
root.title('SysItamara 2024 v1.0')
root.geometry('1280x720')

# Criar o widget Frame para o formulário
form_frame = tk.Frame(root)
form_frame.pack(side=tk.LEFT)

itamara_label = tk.Label(form_frame, text='SysItamara', font= 'Arial 25')
itamara_label.grid(row=0, column=1)

# Criar os widgets Label e Entry para Quarto
quarto_label = tk.Label(form_frame, text='Quarto:')
quarto_label.grid(row=1, column=0, padx=5, pady=5)
quarto_entry = tk.Entry(form_frame)
quarto_entry.grid(row=1, column=1, padx=5, pady=5)

# Criar os widgets Label e Entry para Camareira
camareira_label = tk.Label(form_frame, text='Camareira:')
camareira_label.grid(row=2, column=0, padx=5, pady=5)
camareira_entry = tk.Entry(form_frame)
camareira_entry.grid(row=2, column=1, padx=5, pady=5)

# Criar os widgets Label e Entry para Itens Consumidos
itens_consumidos_label = tk.Label(form_frame, text='Itens Consumidos:')
itens_consumidos_label.grid(row=3, column=0, padx=5, pady=5)
itens_consumidos_entry = tk.Entry(form_frame)
itens_consumidos_entry.grid(row=3, column=1, padx=5, pady=5)

def add_to_spreadsheet():
    # Ler as informações do formulário
    quarto = quarto_entry.get()
    camareira = camareira_entry.get()
    itens_consumidos = itens_consumidos_entry.get()
    data = [quarto, camareira, itens_consumidos]
    
    # Adicionar as informações à planilha
    write_xlsx(data)
    
    # Limpar os campos do formulário
    quarto_entry.delete(0, tk.END)
    camareira_entry.delete(0, tk.END)
    itens_consumidos_entry.delete(0, tk.END)

# Criar o widget Button para adicionar as informações à planilha
add_button = tk.Button(root, text='Adicionar', command=atualizar_table_frame)
add_button.pack(side=tk.BOTTOM)




#df = openpyxl.load_workbook('vendas.xlsx')
#df = pd.read_excel('vendas.xlsx', usecols=range(6))
df = pd.read_excel('vendas.xlsx')
                   
table_frame = tk.Frame(root)
table_frame.pack(side=tk.RIGHT)

tv = tk.ttk.Treeview(table_frame, columns=tuple(df.columns), show='headings')
tv.pack()

for row in df.itertuples(index=False):
    tv.insert('', 'end', values=tuple(row))


# Iniciar o loop principal da janela
root.mainloop()